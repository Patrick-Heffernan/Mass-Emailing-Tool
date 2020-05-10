import openpyxl, ezgmail

wb = openpyxl.load_workbook('Email Addresses.xlsx')
sheet = wb['Sheet1']

emails = {} #Make the excel data into a dictionary
for i in range(1,sheet.max_row-1):
    name = sheet.cell(i, 1).value
    email_address = sheet.cell(i,2).value
    emails[name] = email_address

subject = "Testing!" #Insert subject to your email

for k, v in emails.items(): #Loop through the dictionary data and send email to each address with the corresponding name
   sendTo = v
   content = "Hello, " + str(k) + ". This is a test" #Insert the content of your email
   ezgmail.send(sendTo, subject, content)



